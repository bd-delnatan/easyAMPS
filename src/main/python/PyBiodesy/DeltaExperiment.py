from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
from pathlib import Path
from copy import deepcopy
import matplotlib.pyplot as plt
import numpy as np
import re

# INTERNAL CONSTANTS
ALPHABET = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
# escape special characters translation dictionary
SPECIALTRANS = str.maketrans(
    {
        "+": r"\+",
        "(": r"\(",
        ")": r"\)",
        "$": r"\$",
        ".": r"\.",
        "?": r"\?",
        "[": r"\[",
        "]": r"\]",
    }
)
# channel names from the excel string entry
CHANNELNAMES = {
    "P-SHG": "pSHG",
    "S-SHG": "sSHG",
    "P-FL": "pTPF",
    "S-FL": "sTPF",
}
# end of internal constant


def return_plate_index(well_pos):
    """ this function uses openpyxl's utility function for multi-well grid layout

    It converts letter-based coordinate system to 1-based numeric coordinate
    (e.g. C8 becomes (3,8))
    """
    column, row = coordinate_to_tuple(well_pos)
    return row, column


def niceprint(array, width=100, fmt="{:5.1f}", magnitude=1.0):
    """ print numpy array with formatting and scaling

    Args:
        array(numpy array): array to be printed
        width(int): number of lines before wrapping
        fmt(str): format for floats
        magnitude(float): divisor to the array to display large numbers
    """
    with np.printoptions(linewidth=width):
        print(
            np.array2string(
                array / magnitude, formatter={'float_kind': fmt.format}
            )
        )
        print("x {:10.2f}".format(magnitude))


def get_sheet_prefix(sheetnames):
    """ convenience function to parse the prefix of sheetnames containing data
    """
    ptn = re.compile(r".+\s(Raw|Detail|Summary)")
    base_sheet_prefix = []
    for name in sheetnames:
        strmatch = ptn.match(name)
        if strmatch is not None:
            base_sheet_prefix.append(strmatch.group().split()[0])
    # return only unique names, compress list to set to eliminate repeats
    # then convert back to list
    return list(set(base_sheet_prefix))


class Plate:
    """ Biodesy Delta plate experiment

        This is the primary data structure for handling an experiment
        performed on Biodesy instrument

        Args:
            excel_file(str): path to the result file output by the machine

    """

    def __init__(self, excel_file):
        self.filename = Path(excel_file)
        self._workbook = load_workbook(str(self.filename.resolve()))
        self.Read_Plate = None
        self.relative_data = None

        # for the 'Region' metadata
        self.Ntargetcolumn = 5
        self.Ntransferscolumn = 9
        self.Npolarizercolumn = 2

        # these numbers are specific for 384-well plate
        self.Nrows = 16
        self.Ncolumns = 24
        self.rawreads = None

        # this may be specific to newer Biodesy Delta?
        self._parse_metadata()

        # get experimental detail from 'Region' sheet
        self._parse_region()

        # read data from 'Region Summary', this figures out Nchannels
        self._read_plate_data()

    def __repr__(self):
        """ A string representation of this Plate object

        Prints out global plate information

        """
        strfmt = "Plate name : {:s} (ID {:s})\n"
        strfmt += "Ran : {:s}\n"
        strfmt += "Top-left position : {:s}\n"
        strfmt += "# reads: {:d}, # rows : {:d}, # columns : {:d}\n"
        strfmt += "Instrument : {:s}\n"
        strfmt += "Channels ({:d}): " + "{:s} " * self.Nchannels + "\n"
        return strfmt.format(
            self.name,
            self.readplateID,
            self.datetime,
            self.top_left_pos,
            self.data.Nreads,
            self.height,
            self.width,
            self.instrument,
            self.Nchannels,
            *self.channel_strlist
        )

    def _parse_metadata(self):
        # reads metadata from 'Experiment' sheet
        experiment = self._workbook['Experiment']
        self.name = experiment['B3'].value
        self.readplateID = experiment['B5'].value
        if self.readplateID is None:
            self.readplateID = 'N/A'
        self.project = experiment['B6'].value
        self.platesize = int(experiment['E5'].value)
        self.datetime = experiment['B7'].value.isoformat()
        self.instrument = experiment['E7'].value
        # determine machine version
        version_pattern = re.compile("Delta\s\d+\.\d+\.\d+\.\d+")
        version_strmatch = (
            version_pattern.match(self.instrument)
            .group()
            .strip("Delta ")
            .split(".")
        )
        self.deltaversion = int(version_strmatch[0])

        if self.deltaversion == 3:
            self.platemap_column_offset = 4
            self.region_row_start = 6
            self.region_summary_row_offset = 2
        elif self.deltaversion == 4:
            self.platemap_column_offset = 3
            self.region_row_start = 9
            self.region_summary_row_offset = 3
        else:
            raise NotImplementedError(
                "This Delta version ({:d}) is not supported.".format(
                    self.deltaversion
                )
            )

        self.sheet_prefix = get_sheet_prefix(self._workbook.sheetnames)

        if len(self.sheet_prefix) > 1:
            raise NotImplementedError(
                "Program currently only supports "
                "one experiment read per file."
            )

    def _parse_region(self):
        # get the "Region" sheet
        region_sheet = self.sheet_prefix[0]
        region = self._workbook[region_sheet]

        block_data = [
            cell.value
            for row in region
            for cell in row
            if cell.value is not None
        ]

        # get the indices of where these meta tags
        # plus one because the metadata is the next column over
        target_id = block_data.index("Target:") + 1
        transfer_id = block_data.index("Transfers:") + 1

        # metadata under 'Target'
        target_metadata = dict(
            [
                d
                for d in [
                    block_data[
                        target_id
                        + n : target_id
                        + 2 * self.Ntargetcolumn : self.Ntargetcolumn
                    ]
                    for n in range(self.Ntargetcolumn)
                ]
            ]
        )

        # metadata under 'Transfers'
        transfers_metadata = dict(
            [
                (d[0], d[1:])
                for d in [
                    block_data[transfer_id + n :: self.Ntransferscolumn]
                    for n in range(self.Ntransferscolumn)
                ]
            ]
        )

        if self.deltaversion == 4:
            polarizer_id = block_data.index("Polarizer:") + 1

            # metadata under 'Polarization'
            polarizer_metadata = dict(
                [
                    d
                    for d in [
                        block_data[
                            polarizer_id
                            + n : polarizer_id
                            + 2 * self.Npolarizercolumn : self.Npolarizercolumn
                        ]
                        for n in range(self.Npolarizercolumn)
                    ]
                ]
            )

            self.start_angle = polarizer_metadata['Start Angle']
            self.end_angle = polarizer_metadata['End Angle']

        # assign important information from metadata for read plate
        self.top_left_pos = target_metadata['Top-Left Position']
        self.width = int(target_metadata["Width"])
        self.height = int(target_metadata["Height"])
        self.init_vol = float(target_metadata["Initial Volume"])

        # the number of "Transfers"
        self.Ntransfers = len(transfers_metadata["Source Plate Position"])
        self.source_init_vol = transfers_metadata["Source Initial Volume"]
        self.source_top_left = transfers_metadata[
            "Source Plate Top-Left Position"
        ]
        self.Nbaseline = transfers_metadata["Number of Baseline Reads"]
        self.dt_baseline = transfers_metadata[
            "Baseline Read Interval (seconds)"
        ]
        self.mix_cycles = transfers_metadata["Mix Cycles"]
        self.inject_vol = transfers_metadata["Injection Volume"]
        self.Npostinject = transfers_metadata["Number of Post-Injection Reads"]
        self.dt_postinject = transfers_metadata[
            "Post-Inject Read Interval (seconds)"
        ]

        if self.Ntransfers == 1:
            N = self.Ntransfers - 1
            self.Nbaseline = int(self.Nbaseline[N])
            self.dt_baseline = self.dt_baseline[N]
            self.mix_cycles = self.mix_cycles[N]
            self.inject_vol = self.inject_vol[N]
            self.Npostinject = int(self.Npostinject[N])
            self.dt_postinject = self.dt_postinject[N]
        else:
            # for now, this is a hack? verify this between Delta 3 & 4
            # the number of reads is the sum of all post injection reads in v3
            self.Npostinject = int(sum(self.Npostinject))
            self.Nbaseline = int(sum(self.Nbaseline))
            self.dt_postinject = max(self.dt_postinject)
            self.inject_vol = max(self.inject_vol)

        # account for the start position of plate reads
        start_row, start_column = return_plate_index(self.top_left_pos)
        start_row -= 1
        start_column -= 1
        self.row_labels = [ALPHABET[start_row + i] for i in range(self.height)]
        self.col_labels = [start_column + i + 1 for i in range(self.width)]

    def _read_plate_data(self):
        # reads plate data from 'Region Summary' sheet
        # and compute percent change if there are multiple reads
        # percent change is defined by the change of the last reading to
        # the baseline read
        Nreads = self.Nbaseline + self.Npostinject
        start_row, start_column = return_plate_index(self.top_left_pos)
        # account for corner padding in Excel
        start_row += self.region_summary_row_offset
        start_column += 3

        # read data from "Region Summary"
        summary_sheet = self.sheet_prefix[0] + " Summary"
        region_summary = self._workbook[summary_sheet]

        if self.deltaversion > 3:
            # look for channel names here (these should be fixed positions)
            recorded_channels = [
                region_summary.cell(row=1, column=c).value
                for c in [4, 29, 54, 79]
            ]

            self.channel_strlist = [
                CHANNELNAMES[s] for s in recorded_channels if s is not None
            ]

            self.Nchannels = len(self.channel_strlist)

        else:
            self.channel_strlist = ['pSHG']
            self.Nchannels = 1

        self.platedata = np.zeros(
            (Nreads, self.height, self.width, self.Nchannels)
        )

        for n in range(Nreads):
            row_offset = int(n * (1 + self.Nrows))
            row_begin = start_row + row_offset
            for c in range(self.Nchannels):
                column_offset = int(c * (1 + self.Ncolumns))
                column_begin = start_column + column_offset
                for r, row in enumerate(
                    region_summary.iter_rows(
                        min_row=row_begin,
                        max_row=row_begin + self.height - 1,
                        min_col=column_begin,
                        max_col=column_begin + self.width - 1,
                        values_only=True,
                    )
                ):
                    self.platedata[n, r, :, c] = np.array(row)

        self.data = BiodesyArray(self.platedata, self.channel_strlist)

    def _read_raw_data(self):
        """ parse and store the raw data scans

        Useful for diagnosing the plate "health" of each experiment

        Each plate scans are stored in an dictionary of NumPy arrays.
        self.rawreads['<channel_name>'] contains an 4-D array with shape:
        Nreads x Nrows x Ncolumns x Nscans

        Nscans is the number of scans done by the machine and it is fixed
        to 11 scans. The data that we get in 'Summary' is the median of
        those scans. Nreads also include the 'B1' or baseline scan in addition
        to the number of scans post-injections.

        """
        # these are positions on the excel sheets
        if self.deltaversion == 3:
            # in delta version 3, the raw data resides in 'Detail'
            region_raw = self._workbook[self.sheet_prefix[0] + " Detail"]
            fieldnames = "pSHG pPD"
            raw_read_bounds = dict(
                zip(fieldnames.split(), ((20, 30), (31, 41)))
            )

        elif self.deltaversion == 4:
            # while in delta version 4, the raw data resides in 'Raw'
            region_raw = self._workbook[self.sheet_prefix[0] + " Raw"]

            # for the field names for the "Raw" data channels
            fieldnames = ["pSHG", "sSHG", "pPD", "sPD", "pFL", "sFL"]
            colbounds = [
                (21, 31),
                (39, 49),
                (53, 63),
                (67, 77),
                (85, 95),
                (103, 113),
            ]

            incr = int(4 / self.Nchannels)

            raw_read_bounds = dict(
                zip(
                    fieldnames[::incr], colbounds[::incr]
                )
            )

        # store the raw data in a dictionary
        self.rawreads = {}
        # calculate the row-wise block for each well, there will be Nreads row
        Nreads = self.Nbaseline + self.Npostinject

        # the starting row appears to be the same between Delta 3 & 4
        row_start = 21

        if self.deltaversion ==4 :
            if self.inject_vol > 0.0:
                # account for row-skipping due to injection
                skip = 1
                # number of apparent reads increments by skip amount (as extra
                # empty row)
                Nreads += 1
            else:
                # read all rows, setting this below zero bypasses skipping
                # because skips are checked with she conditional statement
                # if (skip != t, for t in range(N); so normally, t >= 0)
                # having skip=1 will skip at index t=1 (e.g. it will read t=0, and
                # t=2, t=3, t=... so on)
                skip = -1

        if self.deltaversion == 3:
            # figure out if there's a skip due to injection
            # in v3, it happens when there are more than 1 transfers
            if self.Ntransfers > 1 or self.inject_vol>0.0:
                skip = 1
                Nreads += 1
            else:
                skip = -1

        # read the entire raw array, first by channel
        for raw_channel, (mincol, maxcol) in raw_read_bounds.items():
            # each array for raw data is Nreads x Height x Width x 11
            # 11 is the number of scans from each well (machine-specified)
            _dump = np.zeros((Nreads, self.height, self.width, 11))
            channel_is_empty = False
            for n in range(self.width * self.height):
                # the 2D coordinates (row and column)
                _r = n // self.width
                _c = n % self.width
                row_begin = row_start + int(n * Nreads)
                row_end = row_begin + Nreads - 1

                for t, rowdata in enumerate(
                    region_raw.iter_rows(
                        min_row=row_begin,
                        max_row=row_end,
                        min_col=mincol,
                        max_col=maxcol,
                        values_only=True,
                    )
                ):
                    # skip the t-th row of there was injection (0-indexing)
                    # otherwise this is -1, which always evaluates to True
                    if t != skip:
                        # put row-strip as array into working array
                        if rowdata[0] is not None:
                            _dump[t, _r, _c, :] = np.array(rowdata)
                        else:
                            channel_is_empty = True
                            break
            if not channel_is_empty:
                self.rawreads[raw_channel] = _dump

    def _calculate_percent_change(self):
        """ computes relative signal change in all channels from the first read

        If the desired baseline is read as a separate result file, then that
        file needs to be prepended to the current Plate. Use the
        _merge_baseline_read(Plate) method to do this

        """

        # do a deep copy to clone the data structure
        self.relative_data = deepcopy(self.data)

        # then for each channel do the %change calculation
        # using the formula 100 * (data - data(t=0)) / data(t=0)
        for ch_name, ch_data in self.data:
            self.relative_data.__dict__[ch_name] = 100.0 * (
                (
                    self.data.__dict__[ch_name]
                    - self.data.__dict__[ch_name][0, ...]
                )
                / self.data.__dict__[ch_name][0, ...]
            )

    def _parse_plate_legend(self, plate_name):
        """ internal class method to parse the Biodesy plate legend

        Assigns the plate legend to the same matrix layout as the data. This
        function will assign a new attribute to the object with the same name
        as the 'plate_name'. All spaces are replaced with an underscore.

        Args:
            plate_name(str): the Sheet name contained in current Excel worksheet

        """

        # remove and replace spaces with underscores to assign object attribute
        readplate = self._workbook[plate_name]
        attrname = plate_name.replace(" ", "_")

        # get the starting read position
        start_row_index, well_column_number = return_plate_index(
            self.top_left_pos
        )

        # account for padding along rows
        start_row = (
            (start_row_index - 1) * self.Ncolumns + well_column_number + 5
        )
        column_offset = self.platemap_column_offset

        # pre-allocate an empty container array
        platelegend = np.empty((self.height, self.width, 3), dtype='object')

        # change the 'max_col' parameter to get more columns from the metadata
        for r in range(self.height):
            start_row = (
                ((start_row_index + r) - 1) * self.Ncolumns
                + well_column_number
                + 5
            )
            end_row = start_row + self.width - 1
            for i, annotation in enumerate(
                readplate.iter_cols(
                    min_row=start_row,
                    max_row=end_row,
                    min_col=column_offset,
                    max_col=column_offset + 2,
                    values_only=True,
                )
            ):
                # strip trailing characters incase of un-intended space added
                platelegend[r, :, i] = annotation

        self.__setattr__(attrname, PlateLegend(platelegend))

        empty_annotation = []
        # check which annotation is absent
        if np.all(platelegend[:, :, 0] == ''):
            # Substance is empty
            empty_annotation.append('substance')
            print("Substances not found")

        if np.all(platelegend[:, :, 1] == ''):
            # Concentrations is empty
            empty_annotation.append('concentration')
            print("Concentrations not found")

        if np.all(platelegend[:, :, 2] == ''):
            # unit is empty
            empty_annotation.append('unit')
            print("Units not found")
        # assign plate legend to object attribute
        # convert the plate legend as unicode char, max is 256 characters

    def _uniform_read_plate(
        self,
        conjugate_name="labeled protein",
        conjugate_conc=1.0,
        conjugate_unit="uM",
        attrname="Read_Plate",
    ):
        """ assign a uniform conjugate population to read plate

        Args:
            conjugate_name(str): name of conjugate
            conjugate_conc(float): concentration of conjugate in every well
            conjugate_unit(str): concentration unit for conjugate
            attrname(str): name of plate map assigned to Plate. Default is
                "Read_Plate". This shouldn't be changed because it will break
                the convention.
        """
        platelegend = np.empty((self.height, self.width, 3), dtype='object')
        platelegend[:, :, 0] = conjugate_name
        platelegend[:, :, 1] = conjugate_conc
        platelegend[:, :, 2] = conjugate_unit
        self.__setattr__(attrname, PlateLegend(platelegend))

    def _import_plate_legend(self, excel_file, corner_position):
        """ assign an custom plate map to the current plate (Excel file)

        Unique map value/description of each well has their indices grouped
        together. This assigns a 'custom_map' to the current object

        Args:
            excel_file(str): the file path to the plate map
            corner_cell(str): the coordinate for corner cell in Excel format
                (e.g. 'A12')
        """

        xl_filepath = Path(excel_file)
        plate_xlsheet = load_workbook(str(xl_filepath.resolve()))

        start_row, start_column = coordinate_to_tuple(corner_position)
        # offset 1 from corner to skip the header values
        start_row += 1
        start_column += 1

        platelegend = np.empty((self.height, self.width), dtype='object')

        for r, annotation in enumerate(
            plate_xlsheet.active.iter_rows(
                min_col=start_column,
                max_col=start_column + self.width - 1,
                min_row=start_row,
                max_row=start_row + self.height - 1,
                values_only=True,
            )
        ):
            platelegend[r, :] = annotation

        self.__setattr__('custom_map', PlateLegend(platelegend, custom=True))

    def _fetch_unique_experiments(
        self, return_by_group=False, group_by_conjugates=True
    ):
        """ Returns a list of [conjugate, compound] pair that was read

        Every conjugate concention is treated as a different conjugate (e.g.
        will have its own label with their respective concentration).

        Args:
            return_by_group(bool): if True, the unique experiments are grouped
                by its unique group (either compound or conjugate). If False,
                a list containing pairs of unique conjugate-compound are
                returned.
            group_by_conjugates(bool): if True, the each unique conjugate is
                used to group experiments (e.g. per unique conjugate, multiple
                compounds).

        """
        all_conjugates = np.unique(self.Read_Plate.substance)
        all_compounds = np.unique(self.Source_Plate_1.substance)

        expt_pair = []

        for conjugate in all_conjugates:
            conjugate_plate_mask = self.Read_Plate.lookup(conjugate)
            for compound in all_compounds:
                source_plate_mask = self.Source_Plate_1.lookup(compound)
                unique_mask = conjugate_plate_mask & source_plate_mask
                if unique_mask.sum() > 0:
                    expt_pair.append([conjugate, compound])

        if return_by_group:
            grouped_experiments = {}
            if group_by_conjugates:
                for conjugate, compound in expt_pair:
                    if conjugate not in grouped_experiments:
                        grouped_experiments[conjugate] = [compound]
                    elif conjugate in grouped_experiments:
                        grouped_experiments[conjugate].append(compound)
            else:
                for conjugate, compound in expt_pair:
                    if compound not in grouped_experiments:
                        grouped_experiments[compound] = [conjugate]
                    elif compound in grouped_experiments:
                        grouped_experiments[compound].append(conjugate)

            return grouped_experiments
        else:
            return expt_pair

    def unique_experiments(self, relative_data=False):
        """ A data generator for [conjugate, compound] pair

        Generates the associated data structure (Python dict) for the
        (conjugate, compound) pair. The shape of the data is
        Nreads x Nreplicates x Nconcs (Nconcs for the conjugate). Each conjugate
        concentration is treated as a unique conjugate. This generator uses the
        method :func:`PyBiodesy.DeltaExperiment.Plate.get_unique_experiment`

        yields data, (conjugate, compound)

        Args:


        """
        all_conjugates = np.unique(self.Read_Plate.substance)
        all_compounds = np.unique(self.Source_Plate_1.substance)

        expt_pair = []

        for conjugate in all_conjugates:
            conjugate_plate_mask = self.Read_Plate.lookup(conjugate)
            for compound in all_compounds:
                source_plate_mask = self.Source_Plate_1.lookup(compound)
                unique_mask = conjugate_plate_mask & source_plate_mask
                if unique_mask.sum() > 0:
                    data = self.get_unique_experiment(
                        conjugate, compound, relative_data=relative_data
                    )
                    yield data, (conjugate, compound)

    def get_well_identity(self, well_position):
        """ returns the information about a particular well
        """
        row, column = return_plate_index(well_position)
        # python uses 0-indexing, but the well starts at 1. So need to -1
        row -= 1
        column -= 1
        if self.Read_Plate is not None:
            conjugate_info = dict(
                conjugate=self.Read_Plate.substance[row, column],
                concentration=self.Read_Plate.concentration[row, column],
                unit=self.Read_Plate.unit[row, column],
            )
        else:
            conjugate_info = None

        if self.Source_Plate_1 is not None:
            compound_info = dict(
                compound=self.Source_Plate_1.substance[row, column],
                concentration=self.Source_Plate_1.concentration[row, column],
                unit=self.Source_Plate_1.unit[row, column],
            )
        else:
            compound_info = None

        return conjugate_info, compound_info

    def get_unique_experiment(
        self, conjugate_name, compound_name, relative_data=False
    ):
        """ returns a dictionary of unique experiments for a given conjugate+compound

        The output dictionary is grouped by conjugate (from Read Plate) and its
        concentration. Returns None if such combination of experiment does not
        exist.

        The structure of the output is ::

            Conjugate(at a given concentration)
                ├─Channel#1
                │      ├─'concentrations' (Nconcs)
                │      ├─'unit' (str): indicating units
                │      └─'data' (Nreads x Nreplicates x Nconcs)
                │
                ├─Channel#2
                │       ┊
                ┊

        With each 'data' sub-array having the shape
        Nreads x Nreplicate x Nconcentrations

        """

        if relative_data:
            if self.relative_data is None:
                self._calculate_percent_change()
            plate_data = self.relative_data
        else:
            plate_data = self.data

        conjugate_plate_mask = self.Read_Plate.lookup(conjugate_name)
        source_plate_mask = self.Source_Plate_1.lookup(compound_name)

        # check to see if such conjugate + compound pair exists:
        unique_mask = conjugate_plate_mask & source_plate_mask
        exists = unique_mask.sum() > 0

        if exists:
            # figure out how many conjugate concentration was used for this
            # particular compount
            all_conjugate_concs = self.Read_Plate.concentration[unique_mask]
            all_conjugate_units = self.Read_Plate.unit[unique_mask]
            conjugate_concs = np.unique(all_conjugate_concs)
            Nconcs_conjugate = len(conjugate_concs)

            # define output dictionary for 'conjugate_data'
            conjugate_dict = {}

            # go through each unique conjugate concentrations
            for conj_conc in conjugate_concs:
                # form a smaller subset of mask for each conjugate concentration
                unique_submask = unique_mask & (
                    self.Read_Plate.concentration == conj_conc
                )

                # get units for this particular conjugate concentrations
                conj_unit = np.unique(self.Read_Plate.unit[unique_submask])[0]
                compound_unit = np.unique(
                    self.Source_Plate_1.unit[unique_submask]
                )[0]

                # form dictionary key for "Conjugate (10.0 uM)"
                # we can deal with differing precision digits and units here
                conjugate_key = "{:s} ({:0.1f} {:s})".format(
                    conjugate_name, conj_conc, conj_unit
                )

                # now go through each concentrations in the compound
                all_compound_concs = self.Source_Plate_1.concentration[
                    unique_submask
                ]
                compound_concs = np.unique(all_compound_concs)
                Ncompound_concs = int(len(compound_concs))
                Nreplicate = int(len(all_compound_concs) / Ncompound_concs)

                # Now we prepare output data array with the shape:
                # Nreads x Nreplicate x  Nconcs
                #                      (compound)
                channel_dict = {}
                for ch_name, ch_data in plate_data:
                    # pre-allocate array
                    conjugate_data = np.zeros(
                        (self.data.Nreads, Nreplicate, Ncompound_concs)
                    )
                    for i, conc in enumerate(compound_concs):
                        conc_submask = unique_submask & (
                            self.Source_Plate_1.concentration == conc
                        )
                        conjugate_data[:, :, i] = ch_data[:, conc_submask]
                    channel_dict[ch_name] = {
                        'concentrations': compound_concs,
                        'unit': compound_unit,
                        'data': conjugate_data,
                    }
                    conjugate_dict[conjugate_key] = channel_dict

            return conjugate_dict

        else:

            return None

    def merge_baseline_read(self, BaselinePlate):
        """ join off-line baseline read to current plate read

        This method prepends an entire baseline read before the first
        time point. The BiodesyArray handles all the channel matching
        so they don't have to have the same number of channels
        """
        self.data._prepend_array(BaselinePlate.data)

    def append_data(self, Plate):
        """ append a separate data from a different Plate """
        self.data._append_array(Plate.data)

    def get_well_data(self, well_position, relative_data=False):
        row, column = return_plate_index(well_position)
        row_offset, column_offset = return_plate_index(self.top_left_pos)
        row_id = row - row_offset
        col_id = column - column_offset

        if relative_data:
            if self.relative_data is not None:
                data_ = self.relative_data
            else:
                print("Computing relative data ...")
                self._calculate_percent_change()
                data_ = self.relative_data
        else:
            data_ = self.data

        if row_id < self.data.Nrows and col_id < self.data.Ncolumns:
            return {ch_name: ch_data[:, row_id, col_id]
                    for ch_name, ch_data in data_}
        else:
            return None

    def visualize_plate_data(
        self,
        array,
        title="Plate view",
        axis_handle=None,
        colorbar_orientation="vertical",
        **kwargs
    ):
        """ do the augmented matplotlib.pyplot.imshow for 384-well plate

        The extra arguments are passed onto pyplot.imshow(). This should be
        re-factored as a class method because it could be used for any plate
        data. The only thing that needs to be handled is the experiment-specific
        plate boundaries (self.height and self.width parameters)

        """
        if axis_handle is None:
            _view = plt.imshow
            _title = plt.title
            _hlines = plt.hlines
            _vlines = plt.vlines
            _yticks = lambda y, ylabels, fontdict: plt.yticks(
                y, ylabels, **fontdict
            )
            _xticks = lambda x, xlabels, fontdict: plt.xticks(
                x, xlabels, **fontdict
            )
        else:
            _view = axis_handle.imshow
            _title = axis_handle.set_title
            _hlines = axis_handle.hlines
            _vlines = axis_handle.vlines
            _xticks = lambda x, xlabels, **kwargs: [
                axis_handle.set_xticks(x),
                axis_handle.set_xticklabels(xlabels, **kwargs),
            ]
            _yticks = lambda y, ylabels, **kwargs: [
                axis_handle.set_yticks(y),
                axis_handle.set_yticklabels(ylabels, **kwargs),
            ]

        mappable = _view(array, **kwargs)
        _title(title)
        _hlines(
            y=np.arange(1 + self.height) - 0.5,
            xmin=-0.5,
            xmax=self.width - 0.5,
            colors="#999999",
        )
        _vlines(
            x=np.arange(1 + self.width) - 0.5,
            ymin=-0.5,
            ymax=self.height - 0.5,
            colors="#999999",
        )
        _yticks(
            np.arange(len(self.row_labels)),
            self.row_labels,
            fontdict={'fontsize': 'x-small'},
        )
        _xticks(
            np.arange(len(self.col_labels)),
            self.col_labels,
            fontdict={'fontsize': 'x-small'},
        )

        if axis_handle is None:
            plt.colorbar(orientation=colorbar_orientation, pad=0.15)
        else:
            plt.colorbar(
                mappable,
                orientation=colorbar_orientation,
                pad=0.15,
                ax=axis_handle,
            )

    def fit_phase_difference(
        self,
        series_key,
        name="phase_det",
        average_timepoints=True,
        report=True,
    ):
        """ compute phase difference from the "Read Plate" specified via 'series_key'

        This method does a quadratic fit to the TPF vs SHG signal.
        The wells pointed to by 'series_key' should contain varying ratios of
        labeled:unlabeled protein to modulate the SHG and TPF signal. The result
        is the current plate will have a new attribute of PhaseDetermination
        object with the given name. One can use the method from
        PhaseDetermination.correct_background to do SHG signal background correction

        Args:
            series_key(list of str): a list of well labels from the 'Read Plate'
                 that denotes the concentration series of labeled:unlabeled protein
            name(str): the name of experiment or condition. This will be assigned
                to a Plate attribute.
            average_timepoints(bool): if True, the phase difference estimation
                will average across all timepoints

        """

        if self.Read_Plate is not None:
            plate_map = self.Read_Plate.substance
        elif self.custom_map is not None:
            plate_map = self.custom_map.substance
        else:
            assert (self.Read_Plate is not None) and (
                self.custom_map is not None
            ), "No plate maps have assigned"

        Nreads = self.Nbaseline + self.Npostinject
        Nseries = len(series_key)

        # assign the data for X and Y axes (TPF vs SHG signals, in P and S-polarization)
        if average_timepoints:
            # for P-polarization data
            P_x = np.zeros(Nseries)
            P_x_err = np.zeros(Nseries)
            P_y = np.zeros(Nseries)
            P_y_err = np.zeros(Nseries)
            # for S-polarizatiodata
            S_x = np.zeros(Nseries)
            S_x_err = np.zeros(Nseries)
            S_y = np.zeros(Nseries)
            S_y_err = np.zeros(Nseries)
            # go through the series
            for i, key in enumerate(series_key):
                well_filter = plate_map == key
                # separate channel data specific to requested 'series_key'
                # dimension of each data.channel is : Nreads, Nrows, Ncolumns
                _pSHG = self.data.pSHG[:, well_filter]
                _sSHG = self.data.sSHG[:, well_filter]
                _pTPF = self.data.pTPF[:, well_filter]
                _sTPF = self.data.sTPF[:, well_filter]

                # do average across all wells and time
                P_y[i], P_y_err[i] = _pSHG.mean(), _pSHG.std()
                S_y[i], S_y_err[i] = _sSHG.mean(), _sSHG.std()
                P_x[i], P_x_err[i] = _pTPF.mean(), _pTPF.std()
                S_x[i], S_x_err[i] = _sTPF.mean(), _sTPF.std()

            phase_det_P = PhaseDetermination(
                P_x, P_y, P_x_err, P_y_err, name=name + " (P)"
            )
            phase_det_S = PhaseDetermination(
                S_x, S_y, S_x_err, S_y_err, name=name + " (S)"
            )

            phase_det_P.run(report=report)
            phase_det_S.run(report=report)

            self.__setattr__(name + "_P", phase_det_P)
            self.__setattr__(name + "_S", phase_det_S)

            print("Phase determination fit now in:")
            print("P-polarization: {:s}".format(name + "_P"))
            print("S-polarization: {:s}".format(name + "_S"))

        else:

            raise NotImplementedError(
                "Keeping timepoint separate is not yet implemented"
            )


class PlateLegend:
    """A class for plate annotation

    It takes 3-D ndarray of datatype 'object' consisting of the 'Substance',
    'concentrations', 'units' of the plate annotation. Each attribute is a
    2-d array that corresponds to the well positions on the plate.

    Attributes:
        substance: array of 'substance' strings
        concentration: array of floats for 'concentrations'
        unit: array of 'unit' strings

    """

    def __init__(self, plate_legend, custom=False):
        """
        Args:
            plate_legend(ndarray): input raw array of type 'object' with
            ndim=3 for substances, concentrations, and units.

        """
        if plate_legend.ndim == 2:
            # then this is a custom plate legend
            self.substance = plate_legend

            # this is a flag to indicate whether the plate map is custom or not
            self.custom = True
            self.annotated = False

        elif plate_legend.ndim == 3:
            self.custom = False
            (self.Nrows, self.Ncolumns, self.Nfields) = plate_legend.shape

            if np.all(plate_legend == ''):
                self.annotated = False
            else:
                self.annotated = True

            # no substance annotation found
            if np.all(plate_legend[:, :, 0] == ''):
                print("Assigning generic 'substance' as substance")
                plate_legend[:, :, 0] = 'substance'
                self.substance = np.core.defchararray.strip(
                    plate_legend[:, :, 0].astype('>U256')
                )
            else:
                self.substance = np.core.defchararray.strip(
                    plate_legend[:, :, 0].astype('>U256')
                )

            try:
                self.concentration = plate_legend[:, :, 1].astype(np.float64)
            except ValueError:
                # otherwise if it's empty, leave it alone
                self.concentration = plate_legend[:, :, 1]

            if (
                np.all(plate_legend[:, :, 1] == np.nan)
                or np.all(plate_legend[:, :, 1] == '')
                or np.all(plate_legend[:, :, 1] == None)
            ):
                plate_legend[:, :, 1] = 1.0
                print("No concentration was found. Assuming 1.0 unit")
                self.concentration = plate_legend[:, :, 1]

            if (
                np.all(plate_legend[:, :, 2] == '')
                or np.all(plate_legend[:, :, 2] == np.nan)
                or np.all(plate_legend[:, :, 2] == None)
            ):
                plate_legend[:, :, 2] = 'uM'
                print("No unit was found. Using micromolar as default.")

            self.unit = plate_legend[:, :, 2]
        else:
            raise NotImplementedError(
                (
                    "The plate has {:d} dimensions. "
                    "This is not yet implemented"
                ).format(plate_legend.ndim)
            )

    def lookup(self, key):
        """ returns a boolean mask for wells that match a given 'key'

        Args:
            key(str): the pattern to search for in plate 'Substance'.

        Returns:
            ndarray: the logical mask indicating match to 'key'

        """
        if self.custom:
            # look for exact match
            ptn = re.compile(key)
            vecmatch = np.vectorize(lambda x: bool(ptn.match(x)))
            return vecmatch(self.substance)
        else:
            return self.substance == key


class BiodesyArray:
    """ A container class for raw Biodesy plate data

    This class separates each channel and assigns each channel to
    its corresponding channel name for convenience. Timepoint data
    is retained as the first axis. The shape of each channel data is:

    Nreads x Nrows x Ncolumns

    You can also iterate through the channels

    Attributes:
        pSHG(ndarray): a 3D array (time x rows x columns) of the P-SHG channel
        sSHG(ndarray): a 3D array (time x rows x columns) of the S-SHG channel
        pTPF(ndarray): a 3D array (time x rows x columns) of the P-TPF channel
        sTPF(ndarray): a 3D array (time x rows x columns) of the S-TPF channel

    """

    def __init__(self, plate_array_data, channel_strlist):
        assert plate_array_data.ndim == 4, "Plate data must be 4-dimensional"
        assert (
            len(channel_strlist) == plate_array_data.shape[-1]
        ), "Number of channels must match the number of channel names"

        # keep the original data in 'raw_array'
        self.raw_array = plate_array_data

        (
            self.Nreads,
            self.Nrows,
            self.Ncolumns,
            self.Nchannels,
        ) = plate_array_data.shape

        self.shape = (self.Nreads, self.Nrows, self.Ncolumns, self.Nchannels)
        self.channel_names = channel_strlist

        # assign each channel to its corresponding data array
        for n, channel_name in enumerate(channel_strlist):
            self.__setattr__(channel_name, plate_array_data[..., n])

    def __iter__(self):
        # iterators will return two things, channel name and its data
        for channel in self.channel_names:
            yield channel, self.__dict__[channel]

    def __repr__(self):
        reprfmt = "An array of plate data with {:d} channels\n"
        reprfmt += "Nreads = {:d}, Nrows = {:d}, Ncolumns = {:d}\n"
        return reprfmt.format(
            self.Nchannels, self.Nreads, self.Nrows, self.Ncolumns
        )

    def _prepend_array(self, biodesy_array):
        """ prepend data array along time axis """

        assert (
            biodesy_array.Nrows == self.Nrows
        ), "Data does not have equal number of rows"
        assert (
            biodesy_array.Ncolumns == self.Ncolumns
        ), "Data does not have equal number of columns"

        # update data shape
        self.Nreads += biodesy_array.Nreads
        self.shape = (self.Nreads, self.Nrows, self.Ncolumns, self.Nchannels)

        # find matching channel names
        matching_channels = list(
            set(self.channel_names) & set(biodesy_array.channel_names)
        )
        # and prepend input BiodesyArray for each matching channel
        for ch in matching_channels:
            self.__setattr__(
                ch,
                np.concatenate(
                    [biodesy_array.__dict__[ch], self.__dict__[ch]], axis=0
                ),
            )

    def _append_array(self, biodesy_array):
        """ append data array along the time axis """

        assert (
            biodesy_array.Nrows == self.Nrows
        ), "Data does not have equal number of rows"
        assert (
            biodesy_array.Ncolumns == self.Ncolumns
        ), "Data does not have equal number of columns"

        # update data shape
        self.Nreads += biodesy_array.Nreads
        self.shape = (self.Nreads, self.Nrows, self.Ncolumns, self.Nchannels)

        # find matching channel names
        matching_channels = list(
            set(self.channel_names) & set(biodesy_array.channel_names)
        )
        # and append input BiodesyArray for each matching channel
        for ch in matching_channels:
            self.__setattr__(
                ch,
                np.concatenate(
                    [self.__dict__[ch], biodesy_array.__dict__[ch]], axis=0
                ),
            )