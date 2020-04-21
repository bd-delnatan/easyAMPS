"""
This module is the 2nd generation of the data parsing module for Delta
experiment. It leverages the Pandas DataFrame data structure.

"""

import pandas as pd
import numpy as np
import pdb
import re
from statistics import mean, stdev
from math import log10
from collections import namedtuple
from openpyxl import load_workbook
from itertools import product

# A lookup table for well-read order (for sequential plotting)
rows = "ABCDEFGHIJKLMNOP"
cols = [i for i in range(1, 25)]
read_order = {}
sort_order = {}

# loop through column-by-column: for each column, go through every row
# usage: df["Read order"] = df["Well coordinates"].apply(lambda x: read_order[x])
# read order: column is read first, A1,B1,C1,...
for j, col in enumerate(cols):
    for i, row in enumerate(rows):
        wc = "{:s}{:d}".format(row, col)
        read_order[wc] = j * 16 + i + 1

# sort order: row is read first, A1,A2,A3,...
for i, col in enumerate(cols):
    for j, row in enumerate(rows):
        wc = "{:s}{:d}".format(row, col)
        sort_order[wc] = j * 16 + i + 1


# For the "non-essential" header entries.
# dictionary for python-friendly header name as a look-up table
# between Header_Name-to-python_attribute
# e.g. 'Buffer/Additive' is not a valid attribute name, so I shorten it to
# 'Buffer'.

HEADER_LOOKUP = {
    "Buffer/Additive": "Buffer",
    "Custom 1": "Custom1",
    "Custom 2": "Custom2",
    "Custom 3": "Custom3",
    "Custom 4": "Custom4",
    "Custom 5": "Custom5",
}

# basic/minimal headers for sample table
BASIC_HEADERS = ["Row", "Column", "Substance", "Concentration", "Units"]

# metadata for well identities should always have these headers
HDR_COL = [
    "Read Row",
    "Read Column",
    "Read Substance",
    "Read Concentration",
    "Read Units",
    "Source Plate",
    "Source Row",
    "Source Column",
    "Source Substance",
    "Source Concentration",
    "Source Units",
    "Protocol Step #",
    "Read #",
]


class Replicate:
    """ A convenient container class to store replicate data

    This class has a repr that for displaying mean ± std (n) when printing
    to screen or as a pandas cell entry. Simple math operations are supported
    and are carried out using the mean value.

    Args:
    list_of_floats(list or iterable): elements of this iterable will be casted
    as 'float' type for computing its statistics (mean and std).

    Attributes:
    mean: mean of replicate
    std: standard deviation of replicate
    n: number of replicate

    Usage:
    A common usage is to use when aggregating data via `.pivot_table()` or
    `.agg()`.

    ::

        # define the replicator that handles replicate or singletons
        def replicator(x):
            if len(x)>1:
                return Replicate(tuple(x.tolist()))
            else:
                return x

        # then just pass this as the 'aggfunc' keyworded-parameter
        pivdf = df.pivot_table(
            index=["category1", "category2"]
            values=["channel"]
            aggufunc=replicator
        )


    """

    def __init__(self, list_of_floats):
        self.values = list_of_floats
        self.basefmt = "{{{:s}}} ± {{{:s}}} ({{:d}})"

    @property
    def mean(self):
        return mean(self.values)

    @property
    def std(self):
        return stdev(self.values)

    @property
    def n(self):
        return len(self.values)

    def __add__(self, other):
        return self.__float__() + other.__float__()

    def __sub__(self, other):
        return self.__float__() - other.__float__()

    def __mul__(self, other):
        return self.__float__() * other.__float__()

    def __truediv__(self, other):
        return self.__float__() / other.__float__()

    def __le__(self, other):
        return self.__float__() <= other.__float__()

    def __ge__(self, other):
        return self.__float__() >= other.__float__()

    def __lt__(self, other):
        return self.__float__() < other.__float__()

    def __gt__(self, other):
        return self.__float__() > other.__float__()

    def __repr__(self):
        """ this repr codes the behavior of what is displayed """
        absmean = abs(self.mean)
        if absmean > 0:
            magnitude = log10(absmean)
        else:
            # if it's zero, then assign magnitude 1
            magnitude = 1
        if abs(magnitude) > 4:
            strfmt = self.basefmt.format(":.4E", ":.4E")
        elif 1 < abs(magnitude) <= 4:
            strfmt = self.basefmt.format(":.0f", ":.0f")
        elif abs(magnitude) <= 1:
            strfmt = self.basefmt.format(":.3f", ":.3f")

        return strfmt.format(self.mean, self.std, self.n)

    def __str__(self):
        return self.__repr__()


def replicator(x):
    if len(x) > 1:
        return Replicate(tuple(x.tolist()))
    else:
        return x


def expand_replicate(x):
    """ function to expand replicate into its own column

    It's meant to expand a single column column within a pandas dataframe.
    This function does not carry over the column name.

    Usage:

    ::

        expanded = df["column1"].apply(expand_replicate)

        # then you'll need to assign a new column for the replicate
        # by forming a new multi-indexed column.

        multicols = [("column1", col) for col in expanded.columns.tolist()]
        expanded.columns = pd.MultiIndex.from_tuples(multicols)


    """

    if isinstance(x, Replicate):
        return pd.Series({f"rep{i+1}": x for i, x in enumerate(x.values)})
    else:
        return pd.Series({f"rep1": x})


def collapse_replicate(x):
    """ function to collapse replicate into its mean value

    It's meant to collapse the Replicate columns into a single number

    Usage:

        collapsed_series = df["Time Elapsed (s)"].apply(collapse_replicate)

    """

    if isinstance(x, Replicate):
        return x.mean
    else:
        return x


def unstack_index_on_top(mdf, level):
    """ turns an index to column, but puts the new column on top

    """

    wrkdf = mdf.unstack(level).swaplevel(axis=1)
    new_multicols, col_indexer = wrkdf.columns.sortlevel(level)
    df_new = wrkdf.copy().iloc[:, col_indexer]
    df_new.columns = new_multicols
    return df_new


def expand_df_replicates(df, skip_column=None):
    """ expand each Replicate object into its own column """

    def _replicate_array(replicate):
        if isinstance(replicate, Replicate):
            return replicate.n
        else:
            return 1

    # form a new vectorized functions
    count_rep = np.vectorize(_replicate_array)

    # form new multi-columns for the expanded dataframe
    # nmax should have the same count as the number of columns
    nmaxrep = count_rep(df).max(axis=0)
    endpoints = np.cumsum(nmaxrep)
    Ncols_expanded = np.sum(nmaxrep)

    # save original column name and add the expanded name, "Replicate"
    column_names = [df.columns.name] + ["Replicate"]
    columns = df.columns.tolist()

    newcols1 = []
    newcols2 = []

    for col, n in zip(columns, nmaxrep):
        # repeat the column labels for the original columns
        newcols1 += [col] * n
        # and also for the replicates
        newcols2 += [f"rep{i+1}" if n > 1 else "rep1" for i in range(n)]

    newcolumns = [newcols1, newcols2]
    if isinstance(df.index, pd.MultiIndex):
        rows = df.index.remove_unused_levels()
    else:
        rows = df.index
    arr = df.values

    colarr = np.empty((df.values.shape[0], Ncols_expanded))
    colarr.fill(np.nan)

    # define a list to collect the expanded array
    # iterate through rows
    for i, row in enumerate(arr):
        # r = tuple()
        for j, col in enumerate(row):
            if isinstance(col, Replicate):
                _begin = endpoints[j - 1] if j > 0 else 0
                _end = _begin + len(col.values)
                colarr[i, _begin:_end] = col.values
            else:
                colarr[i, j * n] = col

    new_multicols = pd.MultiIndex.from_arrays(newcolumns, names=column_names)

    try:
        return pd.DataFrame(colarr, index=rows, columns=new_multicols)
    except ValueError:
        pdb.set_trace()


def get_sheet_prefix(sheetnames):
    """ convenience function to parse the prefix of sheetnames containing data

    Returns the string prefix of sheets

    Args:
        sheetnames(list of str): Workbook.sheetnames from openpyxl

    Returns:
        str
    """
    ptn = re.compile(r".+\s(Raw|Detail|Summary)")
    base_sheet_prefix = []
    for name in sheetnames:
        strmatch = ptn.match(name)
        if strmatch is not None:
            base_sheet_prefix.append(strmatch.group().split()[0])
    # return only unique names, compress list to set to eliminate repeats
    # then convert back to list
    return list(set(base_sheet_prefix))


# emulate switch-case
def formatter(key):
    """ returns a function according to 'key' given

    If 'key' does not exist, pass x unmodified. If the argument passed
    to formatter('Concentration') is empty, also pass unmodified

    """
    fdict = {"Concentration": lambda x: float(x) if x != "" else x}
    # if key does not exist, return x ( do nothing )
    return fdict.get(key, lambda x: x)


# named tuple for sample table entries, used in parse_sample_table()
WellEntry = namedtuple(
    "WellEntry",
    (
        "Row Column Substance Concentration Units "
        "Buffer Custom1 Custom2 Custom3 Custom4 Custom5"
    ),
)


def check_col_hdr(hdr):
    """ convenient callable to be passed to pandas.read_excel """
    return True if hdr in HDR_COL else False


def parse_sample_table(worksheet, start_row=5):
    """ A generic function to parse the sample table

    This function looks for specific header entries from the sample table.
    Namely Row, Column, Substance, Concentration, Units, Buffer, Custom#...
    It returns a dictionary with wellname : namedtuple pair. The idea is to
    have a structure that would allow expression like
    read_plate['A3'].Substance

    """
    sample_table = {}
    empty_table = True
    for row in worksheet.iter_rows(
        min_row=start_row,
        max_row=None,
        min_col=1,
        max_col=None,
        values_only=True,
    ):
        # determine header column indices
        if row[0] == "Row":
            hdr_dict = {}
            # hdr_dict contains 'header': index pair
            for i, hdr in enumerate(row):
                if hdr in HEADER_LOOKUP.keys():
                    hdr_dict[HEADER_LOOKUP[hdr]] = i
                elif hdr in BASIC_HEADERS:
                    hdr_dict[hdr] = i
        else:
            # read the metadata
            # pass keyworded arguments by unpacking a dictionary built via dict
            # comprehension this automatically fetches the index to the right
            # header/key
            entry = WellEntry(
                **{
                    hdr: formatter(hdr)(row[idx])
                    for hdr, idx in hdr_dict.items()
                    if hdr in WellEntry._fields
                }
            )

            wellname = "{:s}{:d}".format(
                row[hdr_dict["Row"]], row[hdr_dict["Column"]]
            )

            # only enter non-empty wells
            if entry.Substance != "":
                # if there's a Substance filled in, this table is not empty
                empty_table = False
                sample_table[wellname] = entry
    if not empty_table:
        return sample_table
    else:
        return None


def read_detail(
    excel_file,
    expt_prefix="Region",
    drop_injection=True,
    label_rows=True,
    drop_empty_columns=True,
    get_instrument=False,
):
    """ Reads the 'Detail' sheet from a Delta experiment result

    This function can read the output Excel file from 'old' and
    'new'/4-channel-capable Delta instruments.

    Args:
        excel_file(str): the name of Result file.
        expt_prefix(str): the prefix for the sheet name within the Excel file.
            This is usually "Region", but since the user can rename it to
            whatever, it remains as a parameter.
        drop_injection(bool): if True, remove rows that had 'Inj' from the
            "Read #" column. Default is True.
        label_rows(bool): if True, each row is labeled by the well position
            from "Read Row" and "Read Column". Default is True.
        drop_empty_columns(bool): if True, any column that is entirely empty
            will be removed. Default is True.

    """

    expt_sheet_name = "{:s} Detail".format(expt_prefix)

    # header check for # channels
    ch_check_ = pd.read_excel(
        excel_file,
        sheet_name=expt_sheet_name,
        skiprows=18,
        header=None,
        nrows=1,
    )

    ch_check_.dropna(axis="columns", inplace=True)

    # the channel names should be on the first row
    ch_ = ch_check_.iloc[0, :]

    # ch_ now contains (column index, channel name) entry
    if ch_.empty:
        # earlier version (SHG-only) does not have channel headers
        # and doesnt contain an experiment "Raw" sheet
        # letter ranges (Excel style) for data columns in v3
        data_cols_id = "N:R"
        col_interest = ["Clock Time", "Elapsed Time (s)", "Median Counts"]
        ch_names = {"Median Counts": "P-SHG"}
    else:
        # for 4-channel versions
        data_cols_id = "O:S,U:Y,AA:AE,AG:AK"
        col_interest = [
            "Clock Time",
            "Elapsed Time (s)",
            "Median",
            "Median.1",
            "Median.2",
            "Median.3",
        ]
        # create a lookup-table for replacing column names
        ch_names = {}
        for i, ch in enumerate(ch_.values):
            if i == 0:
                ch_names["Median"] = ch
            if i > 0:
                ch_names["Median.{:d}".format(i)] = ch

    # for plate metadata, look for basic column headers
    metadata = pd.read_excel(
        excel_file,
        sheet_name=expt_sheet_name,
        skiprows=19,
        usecols=check_col_hdr,
    )

    # different sheet layout of the outputs will use different 'usecols'
    # only pull-out the columns of interest for the data
    median_data = pd.read_excel(
        excel_file,
        sheet_name=expt_sheet_name,
        skiprows=19,
        usecols=data_cols_id,
    ).loc[:, col_interest]

    # rename the median data column to its channel name
    median_data.rename(columns=ch_names, inplace=True)

    # merge the metadata with median data
    merged_data = pd.concat([metadata, median_data], axis=1)

    if label_rows:
        # label each row for broadcasting operations "A1", "A2", "A3", etc.
        choose_cols = ["Read Row", "Read Column"]
        well_labels = [
            "{:s}{:d}".format(*[row[h] for h in choose_cols])
            for index, row in merged_data[choose_cols].iterrows()
        ]

        # assign well position as row labels (e.g. "A1", "A2", ...)
        merged_data.set_index(pd.Series(well_labels), inplace=True)

    try:
        # fill in missing units
        valid_units = merged_data[merged_data["Read #"] == "Inj"].loc[
            :, ["Read Units", "Source Units"]
        ]

        merged_data["Read Units"].fillna(
            valid_units["Read Units"], axis="index", inplace=True
        )

        merged_data["Source Units"].fillna(
            valid_units["Source Units"], axis="index", inplace=True
        )
    except KeyError:
        pass
        # print("File is from the 'old' style Result file. Skipping Read/Source Units")

    if drop_empty_columns:
        merged_data.dropna(axis="columns", how="all", inplace=True)

    # after the missing units have been filled from 'Inj' row
    # we can remove this row since it contains no data
    if drop_injection:
        merged_data = merged_data[merged_data["Read #"] != "Inj"]

    merged_data["Well coordinates"] = merged_data["Read Row"].astype(
        str
    ) + merged_data["Read Column"].astype(str)

    if get_instrument:
        wb = load_workbook(excel_file)
        ws = wb[expt_sheet_name]
        instrumentSN = ws.cell(row=7, column=5).value
        return merged_data, instrumentSN

    else:

        return merged_data


def read_raw(
    excel_file, expt_prefix="Region", compute_CVs=True, label_rows=True
):
    expt_sheet_name = "{:s} Raw".format(expt_prefix)

    ch_check_ = pd.read_excel(
        excel_file,
        sheet_name=expt_sheet_name,
        skiprows=18,
        header=None,
        nrows=1,
    )

    ch_check_.dropna(axis="columns", inplace=True)
    ch_ = ch_check_.iloc[0, :]
    # skip naming the Photodiode entries
    ch_.drop([50, 64], axis=0, inplace=True)

    data_cols_id = "M:AE,AG:AW,AY:BK,BM:BY,CA:CQ,CS:DI"
    col_interest = [
        "Clock Time",
        "Read #",
        "Elapsed Time (s)",
        "Median Counts",
        "Std Dev.",
        "Median Counts.1",
        "Std Dev..1",
        "Median Counts.2",
        "Std Dev..2",
        "Median Counts.3",
        "Std Dev..3",
    ]

    ch_names = {}
    for i, ch in enumerate(ch_.values):
        if i == 0:
            ch_names["Median Counts"] = ch
            ch_names[f"Std Dev."] = f"Std Dev. ({ch})"
        if i > 0:
            ch_names[f"Median Counts.{i}"] = ch
            ch_names[f"Std Dev..{i}"] = f"Std Dev. ({ch})"

    read_positions = pd.read_excel(
        excel_file,
        sheet_name=expt_sheet_name,
        skiprows=19,
        usecols=lambda x: True if x in ["Read Row", "Read Column"] else False,
    )

    read_data = pd.read_excel(
        excel_file,
        sheet_name=expt_sheet_name,
        skiprows=19,
        usecols=data_cols_id,
    )

    read_data_renamed = read_data.loc[:, col_interest].rename(columns=ch_names)

    # get photodiode readings
    pd_reads = read_data.loc[
        :, read_data.columns.str.match("PD[0-9]{1,2}(.[0-9]{1,2})?")
    ]

    # separate P and S photodiode reads
    p_pd = pd_reads.filter(regex=r"^PD[0-9]{1,2}$")
    s_pd = pd_reads.filter(regex=r"PD[0-9]{1,2}\.")

    # combine all of the data, concatenate all columns
    merged_data = pd.concat(
        [
            read_positions,
            read_data_renamed,
            p_pd.agg(["median", "std"], axis=1).rename(
                columns={"median": "P-PD", "std": "Std Dev. (P-PD)"}
            ),
            s_pd.agg(["median", "std"], axis=1).rename(
                columns={"median": "S-PD", "std": "Std Dev. (S-PD)"}
            ),
        ],
        axis=1,
    )

    if label_rows:
        # label each row for broadcasting operations "A1", "A2", "A3", etc.
        choose_cols = ["Read Row", "Read Column"]
        well_labels = [
            "{:s}{:d}".format(*[row[h] for h in choose_cols])
            for index, row in merged_data[choose_cols].iterrows()
        ]

        # assign well position as row labels (e.g. "A1", "A2", ...)
        merged_data.set_index(pd.Series(well_labels), inplace=True)

    if compute_CVs:
        merged_data_with_CV = merged_data.assign(
            **{
                "CV (P-SHG)": lambda x: x["Std Dev. (P-SHG)"] / x["P-SHG"],
                "CV (S-SHG)": lambda x: x["Std Dev. (S-SHG)"] / x["S-SHG"],
                "CV (P-FL)": lambda x: x["Std Dev. (P-FL)"] / x["P-FL"],
                "CV (S-FL)": lambda x: x["Std Dev. (S-FL)"] / x["S-FL"],
                "CV (P-PD)": lambda x: x["Std Dev. (P-PD)"] / x["P-PD"],
                "CV (S-PD)": lambda x: x["Std Dev. (S-PD)"] / x["S-PD"],
            }
        )

        return merged_data_with_CV

    return merged_data


def compute_relative_change(
    df, baseline_df=None, fluorescence=False, baseline_name="Baseline"
):
    """ Compute %-change for SHG channels of a DataFrame

    This operation is done inplace. The input dataframe, df, must only have a
    one unique "Baseline" read annotated under `Read #` column. The quantity
    from this read is used as the baseline for calculating percent change.
    New columns will be added with the prefix "%Δ"- added to the channel name.

    Args:
        df(dataframe): input dataframe, required.
        baseline_df(dataframe): baseline dataframe, optional.
        fluorescence(bool): if True, also compute percent change for
            fluorescence channel.

    Returns:
        Nothing is returned. Operation is done in place.

    """
    fluorescence_columns = ["P-FL", "S-FL"]
    target_columns = ["P-SHG", "S-SHG"]

    if fluorescence:
        target_columns += fluorescence_columns

    wrk_columns = [
        hdr_col for hdr_col in df.columns if hdr_col in target_columns
    ]

    if baseline_df is None:
        baseline = df[df["Read #"] == baseline_name].copy()
    else:
        baseline = baseline_df[baseline_df["Read #"] == baseline_name].copy()

    baseline.index = baseline["Well coordinates"]
    grouped = df.groupby("Well coordinates")

    for col in wrk_columns:
        new_column = "%Δ{:s}".format(col)
        df[new_column] = grouped[col].apply(
            lambda x: 100.0
            * (x - baseline.loc[x.name, col])
            / baseline.loc[x.name, col]
        )


def correct_SHG(
    df,
    background_P,
    background_S,
    phase_difference_P,
    phase_difference_S,
    p_signs,
    s_signs
):
    """ correct SHG signal in-place

    Args:
        df (dataframe): input dataframe containing P-SHG and S-SHG columns
        background_P (float): background SHG value for unlabeled protein
        phase_difference_P (float): phase difference given in radians.

    Usage::
        correct_SHG(df, 6000., 2000., 1.047, 1.047)

    """
    x1 = df["P-SHG"]
    x2 = df["S-SHG"]

    arg1 = x1 / background_P - np.sin(phase_difference_P) ** 2
    arg2 = x2 / background_S - np.sin(phase_difference_S) ** 2

    df["P-SHGcorr"] = (
        background_P * (np.cos(phase_difference_P) + p_signs * np.sqrt(arg1)) ** 2
    )
    df["S-SHGcorr"] = (
        background_S * (np.cos(phase_difference_S) + s_signs * np.sqrt(arg2)) ** 2
    )


def subtract_blank_fluorescence(df, blankdf):
    """ Subtract background fluorescence in-place

    IMPORTANT: blankdf must have "Well coordinates" as the index (e.g. 'A1', ... ).
    Creates new columns with the suffix "corr" (e.g. P-FLcorr, S-FLcorr)

    Args:
        df (DataFrame): input dataframe with "P-FL" or "S-FL" column to be
        corrected. Must not have indices with "Well coordinates" (e.g. df index
        is 0,1,2,...)
        blankdf (DataFrame): input fluorescence blank. Must have indices with
        "Well coordinates" (e.g. blankdf.index is "A1", "A2", ...)

    """

    if "Well coordinates" not in df.columns:
        df["Well coordinates"] = df["Read Row"].astype(str) + df[
            "Read Column"
        ].astype(str)

    # only subtract wells coordinates that exist in input df
    active_wells = df["Well coordinates"].unique()
    wrk_blankdf = blankdf[blankdf.index.isin(active_wells)]

    grouped = df.groupby(["Well coordinates"])

    df["P-FLcorr"] = grouped["P-FL"].apply(
        lambda x: x - wrk_blankdf.loc[x.name, "P-FL"]
    )
    df["S-FLcorr"] = grouped["S-FL"].apply(
        lambda x: x - wrk_blankdf.loc[x.name, "S-FL"]
    )


def read_sample_table(excel_file, sheet_name, start_row=4):
    """ Reads a sample table using Pandas """
    df = pd.read_excel(excel_file, sheet_name=sheet_name, skiprows=start_row)

    # drop un-assigned rows
    df.dropna(subset=["Substance"], inplace=True)
    df.dropna(axis="columns", how="all", inplace=True)

    choose_cols = ["Row", "Column"]

    # form labels for each well, used for well coordinates
    well_labels = [
        "{:s}{:d}".format(*[row[h] for h in choose_cols])
        for index, row in df[choose_cols].iterrows()
    ]

    # label indices with well coordinates
    df.set_index(pd.Series(well_labels), inplace=True)

    # get sheet prefix
    prefix = sheet_name.split()[0]

    # then append it to the column name
    old_column_names = df.columns.tolist()

    new_column_names = [
        "{:s} {:s}".format(prefix, col) for col in old_column_names
    ]
    column_names = dict(zip(old_column_names, new_column_names))

    df.rename(columns=column_names, inplace=True)

    return df


def merge_experiments(dflist, steps=["tethering", "injection", "+imidazole"]):
    """ routine for merging experiments

    """

    Nexpts = len(dflist)
    Nsteps = len(steps)
    if Nsteps != Nexpts:

        for i, step in enumerate(steps):
            print("{:d}) {:s}".format(i, step))

        raise ValueError(
            "The number of experiments must equal the number of steps"
        )

    Nsteps = 0
    for i, (step, df) in enumerate(zip(steps, dflist)):
        df["Step"] = step
        if i == 0:
            wrkdf = df.copy()
            wrkdf.index = wrkdf["Well coordinates"]
            reftime = wrkdf.loc[wrkdf["Read #"] == 0, "Clock Time"].copy()

        elif i > 0:
            # rename the successize "baseline" steps to 1
            df.loc[df["Read #"] == "Baseline", "Read #"] = 0
            df["Read #"] += Nsteps
            # compute time delta to reference (first DataFrame in list)
            wrk_group = df.groupby("Well coordinates")
            wrk_timedelta = wrk_group["Clock Time"].apply(
                lambda x: x - reftime.loc[x.name]
            )
            df["Elapsed Time (s)"] = wrk_timedelta.apply(
                pd.Timedelta.total_seconds
            )

        steps = df["Read #"].unique().tolist()
        Nsteps += len(steps)

    return pd.concat(dflist, axis=0, sort=False)


# create a lookup table of well coordinates (0-index) and its name
ROWLABELS = "ABCDEFGHIJKLMNOP"
COLLABELS = [int(i + 1) for i in range(24)]

WELLCOORDS = {
    "{:s}{:d}".format(row_char, col_char): (i, j)
    for i, row_char in enumerate(ROWLABELS)
    for j, col_char in enumerate(COLLABELS)
}


def convert_to_long_form(df, value_name="Intensity"):
    """ convert 2d table of a plate layout to the long-form """
    wrk = (
        df.reset_index()
        .melt(id_vars=["index"], value_name=value_name)
        .rename(columns={"index": "Read Row", "variable": "Read Column"})
    )

    wrk["Well coordinates"] = wrk["Read Row"].astype(str) + wrk[
        "Read Column"
    ].astype(str)

    return wrk


class Well:
    """an object to represent a well within a sample table

    Args:
        Row(str): row letter (A-P).
        Column(str): column number (1-24) as a string.
        Substance(str): content of well.
        Concentration(str): concentration of content.
        Unit(str): the unit of concentration.

    """

    def __init__(
        self, Row, Column, Substance=None, Concentration=None, Unit=None
    ):
        self.Row = Row
        self.Column = Column
        self.Substance = Substance
        self.Concentration = Concentration
        self.Unit = Unit
        self.selected = False

    def __repr__(self):
        if self.Substance is not None:
            return "{:s}{:d} ({:s})".format(
                self.Row.upper(), self.Column, self.Substance
            )
        else:
            return "{:s}{:d}".format(self.Row.upper(), self.Column)


class SampleTable:
    """ Sample table for scripted manipulation of sample tables

    This sample table is meant to handle the compound sample table in a high
    throughput experiment.

    Sample table has columns:
        Source Row, Source Column, Source Substance, Source Concentration,
        Source Units, Well coordinates

    When calling to_DataFrame(), a new column 'Source Plate ID' is added



    """

    def __init__(self, excel_file=None, data_frame=None, plate_ID=None):

        # initialise 384-well plate
        self.Nrows = 16
        self.Ncolumns = 24
        self.plate = np.empty((self.Nrows, self.Ncolumns), dtype=object)
        self.rowlabels = [s for s in "ABCDEFGHIJKLMNOP"]
        self.collabels = [int(i + 1) for i in range(self.Ncolumns)]
        self.plateID = plate_ID
        self._excel_filename = None

        if excel_file is not None:
            self._excel_filename = excel_file
            self._df = pd.read_excel(self.excel_filename)

        if data_frame is not None:
            self._df = data_frame
            _plateID = data_frame["Plate Id"].unique()
            if len(_plateID) > 1:
                raise ValueError("Plate ID must be unique. Found more than 1.")
            self.plateID = _plateID[0]

        if data_frame is not None and excel_file is not None:
            raise ValueError("You can't pass both excel_file and data_frame.")

        if excel_file is None and data_frame is None:
            # make a blank plate filled with DMSO
            _allwells = [
                [r, c] for (r, c) in product(self.rowlabels, self.collabels)
            ]
            self._df = pd.DataFrame(
                _allwells, columns=["Source Row", "Source Column"]
            )
            self._df["Source Substance"] = "DMSO"
            self._df["Source Concentration"] = 0.0
            self._df["Source Units"] = "µM"
            self._df["Well coordinates"] = self._df["Source Row"].astype(
                str
            ) + self._df["Source Column"].astype(str)

    # form a grid of named tuples with each element as a 'Well' namedtuple
    def populate_wells(self):
        """ call this function after populating the internal dataframe self._df """
        for i, row_char in enumerate(self.rowlabels):
            if row_char in self._df["Source Row"].values:
                select_rows = self._df[self._df["Source Row"] == row_char]
            else:
                select_rows = None
            for j, col_char in enumerate(self.collabels):
                if select_rows is not None:
                    if col_char in select_rows["Source Column"].values:
                        # a row-slice out of the DataFrame
                        seed_well = select_rows.loc[
                            select_rows["Source Column"] == col_char
                        ]

                        self.plate[i, j] = Well(
                            Row=row_char,
                            Column=col_char,
                            Substance=seed_well["Source Substance"].values[0],
                            Concentration=seed_well[
                                "Source Concentration"
                            ].values[0],
                            Unit=seed_well["Source Units"].values[0],
                        )
                    else:
                        self.plate[i, j] = Well(Row=row_char, Column=col_char)
                else:
                    self.plate[i, j] = Well(Row=row_char, Column=col_char)

    def select_wells(self, list_of_wells):
        coords = [WELLCOORDS[well] for well in list_of_wells]
        for row_id, col_id in coords:
            self.plate[row_id, col_id].selected = True
        return coords

    def plate_value(self, property):
        distill = np.vectorize(lambda x: x.__dict__[property])
        return distill(self.plate)

    def put(self, well_position, compound, concentration, unit="uM"):
        _id = self._df["Well coordinates"] == well_position
        self._df.loc[_id, "Source Substance"] = compound
        self._df.loc[_id, "Source Concentration"] = concentration
        self._df.loc[_id, "Source Units"] = unit

    def perform_dilution(
        self, direction, list_of_wells, Nsteps, dilution_factor=0.5
    ):
        """ do serial dilution
        direction can be: 'up', 'down', 'left' and 'right'

        """
        selected_wells_id = self.select_wells(list_of_wells)

        # pass next_row['up'](row_id, n)
        next_row = {
            "up": lambda r, n: r - n - 1,
            "down": lambda r, n: r + n + 1,
            "left": lambda r, n: r,
            "right": lambda r, n: r,
        }

        next_col = {
            "up": lambda c, n: c,
            "down": lambda c, n: c,
            "left": lambda c, n: c - n - 1,
            "right": lambda c, n: c + n + 1,
        }

        # for checking plate boundary limits
        valid_limit = (
            lambda r, c: True if (0 <= r < 16) and (0 <= c < 24) else False
        )

        # go through each selected wells by (row, column) index pair
        for row_id, col_id in selected_wells_id:
            well_ = self.plate[row_id, col_id]
            # if this well has been assigned
            if well_.Substance is not None:
                # choose the next well
                for n in range(Nsteps):
                    next_row_id = next_row[direction](row_id, n)
                    next_col_id = next_col[direction](col_id, n)
                    # do boundary checks
                    if valid_limit(next_row_id, next_col_id):
                        nextwell_ = self.plate[next_row_id, next_col_id]
                    else:
                        # if the next well for this one violates boundary
                        # skip it
                        break

                    # if this well is selected, stop doing dilution steps
                    if nextwell_.selected:
                        break

                    # otherwise dilute, do the dilution step
                    # copy substance info from 'seed' well
                    nextwell_.Substance = "{:s}".format(well_.Substance)
                    nextwell_.Unit = "{:s}".format(well_.Unit)
                    nextwell_.Concentration = well_.Concentration * (
                        dilution_factor ** (n + 1)
                    )

    def to_DataFrame(self):
        # first convert everything to dictionary
        work_dict = {
            "Source Row": [],
            "Source Column": [],
            "Source Substance": [],
            "Source Concentration": [],
            "Source Units": [],
        }
        for (i, j), well in np.ndenumerate(self.plate):
            if well.Substance is not None:
                work_dict["Source Row"].append(well.Row)
                work_dict["Source Column"].append(well.Column)
                work_dict["Source Substance"].append(well.Substance)
                work_dict["Source Concentration"].append(well.Concentration)
                work_dict["Source Units"].append(well.Unit)
        out_df = pd.DataFrame(work_dict)
        # add unique plate ID
        out_df["Source Plate ID"] = self.plateID
        return out_df


class Plate:
    def __init__(self, excel_file):
        self._workbook = load_workbook(excel_file)
        # check if raw sheets exists
        self.raw_exists = (
            sum(
                [
                    1
                    for sheetname in self._workbook.sheetnames
                    if sheetname[:-3] == "Raw"
                ]
            )
            > 0
        )

        self.Read_Plate = parse_sample_table(self._workbook["Read Plate"])
        self.Nsource_plates = 0

        source_plates = [
            sheetname
            for sheetname in self._workbook.sheetnames
            if sheetname.startswith("Source Plate")
        ]

        for src in source_plates:
            source_plate_ = parse_sample_table(self._workbook[src])
            if source_plate_ is not None:
                attr_str = src.replace(" ", "_")
                self.__setattr__(attr_str, source_plate_)
                self.Nsource_plates += 1


class Abbreviate:
    """ context manager for abbreviating long variables """

    def __init__(self, **kwargs):
        self.abbrs = kwargs
        self.store = {}

    def __enter__(self):
        for key, value in self.abbrs.iteritems():
            try:
                self.store[key] = globals()[key]
            except KeyError:
                pass
            globals()[key] = value

    def __exit__(self, *args, **kwargs):
        for key in self.abbrs:
            try:
                globals()[key] = self.store[key]
            except KeyError:
                del globals()[key]


def stretch_xy(x, y, jitter=False):
    """ returns x and y for plotting with replicates

    In cases where there are n-data points per single x (e.g. replicates),
    x is repeated n-times.

    Args:
        x(iterable): x data. An iterable.
        y(iterable): y data. Each element in y is another iterable with
        of length n.

    Returns:
        two tuples of x and y values where x is repeated

    """
    from random import random

    xy = []
    for x_, y_ in zip(x, y):
        if isinstance(y_, np.float64):
            xy.append((x_, y_))
        else:
            if jitter:
                # repeat x with jitter * uniform()
                xrep = (x_ + random() * jitter for _ in range(len(y_)))
            else:
                # repeat x with no jitter
                xrep = (x_,) * len(y_)
            for pair in tuple(zip(xrep, y_)):
                xy.append(pair)
    lx, ly = zip(*xy)
    return lx, ly
