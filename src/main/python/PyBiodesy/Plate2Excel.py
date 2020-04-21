"""
Modules for saving plate maps to Excel
"""
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils.cell import coordinate_to_tuple
import matplotlib.pyplot as plt
import re
import random

# convenient short-hand of functions
clamp = lambda x: max(min(x, 255), 0)
seq = lambda n: [float(i) / float(n) for i in range(n)]


def vec2hex(arr):
    # converts colormap list array from pyplot.cm to hex
    f2i = lambda f: int(f * 255)  # float to rgb uint8
    M, N = arr.shape
    hexlist = []
    for m in range(M):
        vec = arr[m, :]
        rgbi = [clamp(f2i(v)) for v in vec]
        hexlist.append("{0:02X}{1:02X}{2:02X}".format(*rgbi))
    return hexlist


# color map fetcher from matplotlib.pyplot, Qualitative colormaps
Viridis = lambda n: vec2hex(plt.cm.viridis(seq(n))[:, :3])
Rainbow = lambda n: vec2hex(plt.cm.rainbow(seq(n))[:, :3])
Set3 = lambda n: vec2hex(plt.cm.Set3(seq(n))[:, :3])
Set2 = lambda n: vec2hex(plt.cm.Set2(seq(n))[:, :3])
Set1 = lambda n: vec2hex(plt.cm.Set1(seq(n))[:, :3])
Dark2 = lambda n: vec2hex(plt.cm.Dark2(seq(n))[:, :3])

header_keys_v4 = {
    "A5": "Row",
    "B5": "Column",
    "C5": "Substance",
    "D5": "Concentration",
    "E5": "Units",
    "F5": "Buffer/Additive",
    "G5": "Custom 1",
    "H5": "Custom 2",
    "I5": "Custom 3",
    "J5": "Custom 4",
    "K5": "Custom 5",
}

header_keys_v3 = {
    "A5": "Row",
    "B5": "Column",
    "C5": "Well ID",
    "D5": "Substance",
    "E5": "Concentration",
    "F5": "Units",
    "G5": "Buffer/Additive",
    "H5": "Custom 1",
    "I5": "Custom 2",
    "J5": "Custom 3",
    "K5": "Custom 4",
    "L5": "Custom 5",
}

column_widths_v3 = {
    "A": 4.83203125,
    "B": 7.83203125,
    "C": 7.83203125,
    "D": 13.83203125,
    "E": 13.0,
    "F": 10.83203125,
    "G": 14.83203125,
    "H": 11.83203125,
    "I": 16.83203125,
    "J": 22.83203125,
    "K": 30.83203125,
    "L": 90.83203125,
}

column_widths_v4 = {
    "A": 4.83203125,
    "B": 7.83203125,
    "C": 13.83203125,
    "D": 13.0,
    "E": 10.83203125,
    "F": 14.83203125,
    "G": 11.83203125,
    "H": 16.83203125,
    "I": 22.83203125,
    "J": 30.83203125,
    "K": 90.83203125,
}

row_labels = "ABCDEFGHIJKLMNOP"


def return_plate_index(well_pos):
    """ this function uses openpyxl's utility function for multi-well grid layout

    It converts letter-based coordinate system to 1-based numeric coordinate
    (e.g. C8 becomes (3,8))
    """
    column, row = coordinate_to_tuple(well_pos)
    return row, column


BOLDFONT = Font(size=10, name='Arial', bold=True)
BOLDFONT2 = Font(size=12, name='Arial', bold=True)
ITALICFONT = Font(size=12, name='Arial', italic=True)

HIDDENFONT = Font(size=6, name='Arial', color='FFFFFF')

GRAYBG = PatternFill(
    start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"
)

CENTER_ALIGNMENT = Alignment(horizontal="center")
RIGHT_ALIGNMENT = Alignment(horizontal="right")

HAIRDOT_SIDE = Side(border_style='hair', style='thin', color='999999')
SOLID_SIDE = Side(style='thin', color='000000')

CONTENT_BORDER = Border(
    left=HAIRDOT_SIDE,
    right=HAIRDOT_SIDE,
    top=HAIRDOT_SIDE,
    bottom=HAIRDOT_SIDE,
)

SOLID_BORDER = Border(
    left=SOLID_SIDE, right=SOLID_SIDE, top=SOLID_SIDE, bottom=SOLID_SIDE
)

NULL_SIDE = Side(border_style=None)

NO_BORDER = Border(
    left=NULL_SIDE, right=NULL_SIDE, top=NULL_SIDE, bottom=NULL_SIDE
)


class SampleTable:
    """ A class for writing Excel sample table """

    def __init__(self, filename, version=4):
        self.filename = Path(filename)
        self.wb = Workbook()
        self.read_plate = self.wb.create_sheet("Read Plate", index=1)
        self.source_plate = self.wb.create_sheet("Source Plate 1", index=2)
        self.read_plate.sheet_view.showGridLines = False
        self.source_plate.sheet_view.showGridLines = False

        # remove default 'Sheet'
        del self.wb['Sheet']

        self.deltaversion = version
        # handle v3 and v4 software headers for sample table
        self.header_keys = (header_keys_v3, header_keys_v4)[
            self.deltaversion == 4
        ]
        self.column_widths = (column_widths_v3, column_widths_v4)[
            self.deltaversion == 4
        ]

        for pos, value in self.header_keys.items():
            self.read_plate[pos] = value
            self.read_plate[pos].font = BOLDFONT
            self.read_plate[pos].alignment = CENTER_ALIGNMENT
            self.source_plate[pos] = value
            self.source_plate[pos].font = BOLDFONT
            self.source_plate[pos].alignment = CENTER_ALIGNMENT

        if self.deltaversion == 3:
            self.read_plate['A1'].value = "Delta03:r-plate"

            self.source_plate['A1'].value = "Delta03:s-plate"

            self.read_plate.merge_cells("A4:J4")
            self.source_plate.merge_cells("A4:J4")
            header_columns = "ABCDEFGHIJKL"

        elif self.deltaversion == 4:
            
            # Note, 190607 D.E.
            # Software always looks for 'Delta03'!! tried on 
            # Delta 8, v4.0.0.25

            self.read_plate['A1'] = "Delta03:r-plate"
            self.source_plate['A1'] = "Delta03:s-plate"
            self.read_plate.merge_cells("A4:I4")
            self.source_plate.merge_cells("A4:I4")
            header_columns = "ABCDEFGHIJK"

        # hide the first row
        self.read_plate['A1'].font = HIDDENFONT
        self.source_plate['A1'].font = HIDDENFONT
        self.read_plate.row_dimensions[1].height = 7.25
        self.source_plate.row_dimensions[1].height = 7.25

        self.read_plate['A4'].value = 'Sample Table'
        self.read_plate['A4'].font = Font(size=14, name="Arial", bold=True)
        self.read_plate['A4'].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        self.read_plate.row_dimensions[4].height = 27

        for column in header_columns:
            self.read_plate.column_dimensions[
                column
            ].width = self.column_widths[column]
            self.source_plate.column_dimensions[
                column
            ].width = self.column_widths[column]

        self.source_plate['A4'].value = 'Sample Table'
        self.source_plate['A4'].font = Font(size=14, name="Arial", bold=True)
        self.source_plate['A4'].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        self.source_plate.row_dimensions[4].height = 27

        self.fill_metadata()
        self.fill_headers()

    def fill_metadata(self):

        if self.deltaversion == 3:
            self.read_plate.merge_cells("B3:C3")
            self.read_plate.merge_cells("D3:F3")
            self.read_plate['B3'].value = "Labware Type:"
            self.read_plate['B3'].font = BOLDFONT2
            self.read_plate['B3'].alignment = RIGHT_ALIGNMENT
            self.read_plate['D3'].fill = GRAYBG
            self.read_plate['D3'].border = CONTENT_BORDER
            self.read_plate['E3'].border = CONTENT_BORDER
            self.read_plate['F3'].border = CONTENT_BORDER
            self.read_plate['H3'].value = "Plate ID:"
            self.read_plate['H3'].font = BOLDFONT2
            self.read_plate['H3'].alignment = RIGHT_ALIGNMENT
            self.read_plate['I3'].fill = GRAYBG
            self.read_plate['I3'].border = CONTENT_BORDER
            self.source_plate.merge_cells("A3:C3")
            self.source_plate.merge_cells("D3:F3")
            self.source_plate.merge_cells("A2:C2")
            self.source_plate['A3'].value = "Labware Type:"
            self.source_plate['A3'].font = BOLDFONT2
            self.source_plate['A3'].alignment = RIGHT_ALIGNMENT
            self.source_plate['D3'].fill = GRAYBG
            self.source_plate['D3'].border = SOLID_BORDER
            self.source_plate['E3'].border = SOLID_BORDER
            self.source_plate['F3'].border = SOLID_BORDER
            self.source_plate['A2'].value = "Plate Position:"
            self.source_plate['A2'].font = BOLDFONT2
            self.source_plate['A2'].alignment = RIGHT_ALIGNMENT
            self.source_plate['D2'].fill = GRAYBG
            self.source_plate['D2'].border = SOLID_BORDER
            self.source_plate['H3'].value = "Plate ID:"
            self.source_plate['H3'].font = BOLDFONT2
            self.source_plate['H3'].alignment = RIGHT_ALIGNMENT
            self.source_plate['I3'].fill = GRAYBG
            self.source_plate['I3'].border = CONTENT_BORDER

        if self.deltaversion == 4:
            self.read_plate.merge_cells("D3:F3")
            self.read_plate['C3'].value = "Labware Type:"
            self.read_plate['C3'].font = BOLDFONT2
            self.read_plate['C3'].alignment = RIGHT_ALIGNMENT
            self.read_plate['D3'].fill = GRAYBG
            self.read_plate['D3'].border = CONTENT_BORDER
            self.read_plate['E3'].border = CONTENT_BORDER
            self.read_plate['F3'].border = CONTENT_BORDER
            self.source_plate.merge_cells("A2:C2")
            self.source_plate.merge_cells("D3:F3")
            self.source_plate['A2'].value = "Plate Position:"
            self.source_plate['A2'].font = BOLDFONT2
            self.source_plate['A2'].alignment = RIGHT_ALIGNMENT
            self.source_plate['C3'].value = "Labware Type:"
            self.source_plate['C3'].font = BOLDFONT2
            self.source_plate['C3'].alignment = RIGHT_ALIGNMENT
            self.source_plate['D2'].fill = GRAYBG
            self.source_plate['D2'].border = SOLID_BORDER
            self.source_plate['D3'].fill = GRAYBG
            self.source_plate['D3'].border = SOLID_BORDER
            self.source_plate['E3'].border = SOLID_BORDER
            self.source_plate['F3'].border = SOLID_BORDER

    def fill_headers(self):
        startcolumn = (4, 3)[self.deltaversion == 4]
        endcolumn = (13, 12)[self.deltaversion == 4]
        for i in range(6, 390):
            for j in range(startcolumn, endcolumn):
                self.read_plate.cell(row=i, column=j).fill = GRAYBG
                self.read_plate.cell(row=i, column=j).border = CONTENT_BORDER
                self.source_plate.cell(row=i, column=j).fill = GRAYBG
                self.source_plate.cell(row=i, column=j).border = CONTENT_BORDER

        for i, letter in enumerate("ABCDEFGHIJKLMNOP"):
            for col in range(0, 24):
                r = 6 + i + col + 23 * i
                self.read_plate.cell(row=r, column=1).value = letter
                self.read_plate.cell(row=r, column=2).value = col % 24 + 1
                self.read_plate.cell(
                    row=r, column=1
                ).alignment = CENTER_ALIGNMENT
                self.read_plate.cell(
                    row=r, column=2
                ).alignment = CENTER_ALIGNMENT
                self.source_plate.cell(row=r, column=1).value = letter
                self.source_plate.cell(row=r, column=2).value = col % 24 + 1
                self.source_plate.cell(
                    row=r, column=1
                ).alignment = CENTER_ALIGNMENT
                self.source_plate.cell(
                    row=r, column=2
                ).alignment = CENTER_ALIGNMENT

                if self.deltaversion == 3:
                    self.read_plate.cell(
                        row=r, column=3
                    ).value = "{:s}{:d}".format(letter, col % 24 + 1)
                    self.read_plate.cell(
                        row=r, column=3
                    ).alignment = CENTER_ALIGNMENT
                    self.source_plate.cell(
                        row=r, column=3
                    ).value = "{:s}{:d}".format(letter, col % 24 + 1)
                    self.source_plate.cell(
                        row=r, column=3
                    ).alignment = CENTER_ALIGNMENT
                    self.read_plate.cell(
                        row=r, column=3
                    ).border = CONTENT_BORDER
                    self.source_plate.cell(
                        row=r, column=3
                    ).border = CONTENT_BORDER

                self.read_plate.cell(row=r, column=1).border = CONTENT_BORDER
                self.source_plate.cell(row=r, column=1).border = CONTENT_BORDER
                self.read_plate.cell(row=r, column=2).border = CONTENT_BORDER
                self.source_plate.cell(row=r, column=2).border = CONTENT_BORDER

    def annotate_read_plate(self, content_dict):
        """ Fill in annotation along rows given a dictionary """

        # fill column offset
        offset = (0, 1)[self.deltaversion == 3]

        for wellkey, welldata in content_dict.items():
            _row, _column = return_plate_index(wellkey)
            _r = 6 + (_row - 1) + (_column - 1) + 23 * (_row - 1)
            self.read_plate.cell(row=_r, column=3 + offset).value = welldata[
                'substance'
            ]
            self.read_plate.cell(row=_r, column=4 + offset).value = welldata[
                'concentration'
            ]
            self.read_plate.cell(row=_r, column=5 + offset).value = welldata[
                'unit'
            ]

    def annotate_source_plate(self, content_dict):

        # fill column offset
        offset = (0, 1)[self.deltaversion == 3]

        for wellkey, welldata in content_dict.items():
            _row, _column = return_plate_index(wellkey)
            _r = 6 + (_row - 1) + (_column - 1) + 23 * (_row - 1)
            self.source_plate.cell(row=_r, column=3 + offset).value = welldata[
                'substance'
            ]
            self.source_plate.cell(row=_r, column=4 + offset).value = welldata[
                'concentration'
            ]
            self.source_plate.cell(row=_r, column=5 + offset).value = welldata[
                'unit'
            ]

    def close_and_save(self):
        self.wb.save(str(self.filename))
        self.wb.close()


class ReadableSampleTable:
    """ A more natural layout for saving sample table 

    This class is meant to save the grid-layout data format from the GUI to the
    more readable table layout in Excel.

    """
    # a simple lookup table for mapping well position (grid) to table position
    # (grid). The well position is uses 1-indexing

    def __init__(self, filename):
        self.filename = Path(filename)
        self.wb = Workbook()
        self.read_plate = self.wb.create_sheet("Read Plate", index=1)
        self.source_plate = self.wb.create_sheet("Source Plate 1", index=2)
        self.origin = (3,3)
        self.row_gap = 3
        self.Nrows = 16
        self.Ncols = 24
        self.row_labels = "ABCDEFGHIJKLMNOP"
        self.col_labels = [i for i in range(self.Ncols)]
        self.annotations = ['Substance', 'Concentration', 'Unit']
        # remove default 'Sheet'
        del self.wb['Sheet']

    def _label_blocks(self, target_worksheet):
        for n, label in enumerate(self.annotations):
            row_offset = n * (self.Nrows + self.row_gap)

            # label the block
            cell_ = target_worksheet.cell(
                row=row_offset + self.origin[0] - 1,
                column=self.origin[1]
            )
            cell_.value = label
            cell_.fill = GRAYBG
            cell_.alignment = CENTER_ALIGNMENT

            # label rows
            for i, r in enumerate(self.row_labels):
                cell_ = target_worksheet.cell(
                    row=row_offset + self.origin[0] + i + 1,
                    column=self.origin[1]
                )
                cell_.value = r
                cell_.fill = GRAYBG
                cell_.alignment = CENTER_ALIGNMENT

            # label columns
            for c in range(self.Ncols):
                cell_ = target_worksheet.cell(
                    row=row_offset + self.origin[0],
                    column=self.origin[1] + c + 1
                )
                cell_.value = c+1
                cell_.fill = GRAYBG
                cell_.alignment = CENTER_ALIGNMENT

    def _fill_sheet(self, content_dict, target_worksheet, content_color):
        for wellkey, welldata in content_dict.items():
            _row, _column = return_plate_index(wellkey)
            for n, label in enumerate(self.annotations):
                row_offset = n * (self.Nrows + self.row_gap)
                cell_ = target_worksheet.cell(
                    row = row_offset + self.origin[0] + _row,
                    column = self.origin[1] + _column
                )
                cell_.value = welldata[label.lower()]
                fill_color = Color(rgb=content_color[wellkey])
                cell_.fill = PatternFill("solid", fill_color)

    def fill_source_plate(self, content_dict, content_color):
        self._label_blocks(self.source_plate)
        self._fill_sheet(content_dict, self.source_plate, content_color)

    def fill_read_plate(self, content_dict, content_color):
        self._label_blocks(self.read_plate)
        self._fill_sheet(content_dict, self.read_plate, content_color)

    def close_and_save(self):
        self.wb.save(self.filename)
        self.wb.close()

def platemap2table(platemap_file, output_file, assign_color=True, force_version4=True):
    """ Converts a platemap file to a more human-readable table layout """
    # internal helper function
    def fill_sheet(
        target_worksheet,
        source_worksheet,
        colormapper,
        offset=0,
        assign_color=True,
    ):
        annotations = ['Substance', 'Concentration', 'Unit']
        row_gap = 3

        if assign_color:
            # determine the number of unique substances
            substances = []
            # iterate through substance column
            for cell in source_worksheet.iter_rows(
                min_row=6,
                max_row=389,
                min_col=3 + offset,
                max_col=3 + offset,
                values_only=True,
            ):
                if cell is not None and cell[0] != '':
                    substances.append(cell[0])
            unique_substance = list(set(substances))
            # shuffle substance names 
            # so that 'similar' things are not colored the same way
            random.shuffle(unique_substance)
            Nsubstances = len(unique_substance)
            color_list = colormapper(Nsubstances)

            # make 'substance': 'color' mapping
            color_dict = {
                substance: PatternFill(
                    start_color=hexcolor, end_color=hexcolor, fill_type="solid"
                )
                for substance, hexcolor in zip(unique_substance, color_list)
            }

        for n, label in enumerate(annotations):
            # row block offset
            row_offset = n * (Nrows + row_gap)
            # label the block
            cell_ = target_worksheet.cell(
                row=row_offset + origin[0] - 1, column=origin[1]
            )
            cell_.value = label
            cell_.fill = GRAYBG
            cell_.alignment = CENTER_ALIGNMENT

            for i, r in enumerate(row_labels):
                cell_ = target_worksheet.cell(
                    row=row_offset + origin[0] + i + 1, column=origin[1]
                )
                cell_.value = r
                cell_.fill = GRAYBG
                cell_.alignment = CENTER_ALIGNMENT

            for c in range(Ncols):
                cell_ = target_worksheet.cell(
                    row=row_offset + origin[0], column=origin[1] + c + 1
                )
                cell_.value = c + 1
                cell_.fill = GRAYBG
                cell_.alignment = CENTER_ALIGNMENT

        for row, (r, c) in long2grid_lut.items():

            substance = source_worksheet.cell(row=row, column=3 + offset).value
            concentration = source_worksheet.cell(
                row=row, column=4 + offset
            ).value
            unit = source_worksheet.cell(row=row, column=5 + offset).value
            
            #substance block
            target_worksheet.cell(row=r, column=c).value = substance

            #concentration block
            try:
                target_worksheet.cell(row=r + Nrows + row_gap, column=c).value = (
                    float(concentration)
                    if concentration is not None and concentration!=''
                    else concentration
                )

            except ValueError:
                # if  somehow the concentration is given as string?
                target_worksheet.cell(row=r + Nrows + row_gap, column=c).value = (
                    concentration
                    )
            # unit block
            target_worksheet.cell(
                row=r + 2 * (Nrows + row_gap), column=c
            ).value = unit

            if assign_color:
                if substance is not None and substance != '':
                    try:
                        target_worksheet.cell(
                            row=r, column=c
                        ).fill = color_dict[substance]
                    except:
                        print(color_dict)
                        raise

    workbook = load_workbook(platemap_file)

    # check validity of sheetnames
    read_plate_sheet = workbook['Read Plate']
    source_plate_sheet = workbook['Source Plate 1']

    # get delta version
    verstr_ptn = re.compile("Delta[0-9]{2}\:")
    deltaversion = verstr_ptn.match(read_plate_sheet['A1'].value).group()
    deltaversion = int(deltaversion.rstrip(":").strip("Delta"))

    # column offset, delta v3 has an extra column for Well name
    offset = (0, 1)[deltaversion == 3]
    if force_version4:
        offset = 0
    Nrows = 16
    Ncols = 24

    # define 'origin' for the grid
    origin = (3, 3)

    # starting row for 'long' format
    start_row = 6

    # compose a lookup table for going from long2grid
    long2grid_lut = {
        start_row
        + i: (origin[0] + (i // Ncols) + 1, origin[1] + (i % Ncols) + 1)
        for i in range(Nrows * Ncols)
    }

    # having the inverse mapping is done by flipping value:key pair
    # grid2long_lut = {
    #     value: key for key, value in long2grid_lut.items()
    # }

    # create a new file
    new_workbook = Workbook()
    conjugate_sheet = new_workbook.create_sheet("Read Plate")
    compound_sheet = new_workbook.create_sheet("Source Plate")
    del new_workbook['Sheet']

    fill_sheet(conjugate_sheet, read_plate_sheet, Set3, offset=offset)
    fill_sheet(compound_sheet, source_plate_sheet, Set3, offset=offset)

    new_workbook.save(output_file)
    new_workbook.close()


